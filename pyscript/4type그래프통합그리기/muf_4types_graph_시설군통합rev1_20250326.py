import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from datetime import datetime
from collections import defaultdict

# openpyxl을 이용한 엑셀 작성 및 서식
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# -----------------------------------------------------------------------------
# (0) 폰트 & LaTeX 설정
# -----------------------------------------------------------------------------
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['mathtext.default'] = 'regular'

# -----------------------------------------------------------------------------
# (1) 영문 장소명/물질명 매핑
# -----------------------------------------------------------------------------
location_mapping = {
    "가산A1타워주차장": "Underground Parking Facility",
    "에이샵스크린골프": "Indoor Golf Simulation Facility",
    "영통역대합실": "Subway Station",
    "영통역지하역사": "Subway Station",
    "영통역통합": "Subway Station",
    "이든어린이집": "Childcare Center",
    "좋은이웃데이케어센터": "Daycare Center",
    "좋은이웃데이케어센터통합": "Daycare Center",
    "하이씨앤씨학원": "Educational Facility"
}

pollutant_english = {
    "pm10": "PM10",
    "pm25": "PM2.5",
    "co2": "CO2",
    "voc": "VOC",
    "temp": "Temperature",
    "humi": "Humidity",
    "hcho": "HCHO",
    "co": "CO",
    "no2": "NO2"
}

# -----------------------------------------------------------------------------
# (2) 시설유형/기준
# -----------------------------------------------------------------------------
facility_type_mapping = {
    "가산A1타워주차장": "실내주차장",
    "에이샵스크린골프": "실내체육시설",
    "영통역대합실": "다중이용시설",
    "영통역지하역사": "다중이용시설",
    "이든어린이집": "어린이집",
    "좋은이웃데이케어센터": "노인요양시설",
    "좋은이웃데이케어센터1": "노인요양시설",
    "좋은이웃데이케어센터2": "노인요양시설",
    "좋은이웃데이케어센터3": "노인요양시설",
    "좋은이웃데이케어센터4": "노인요양시설",
    "하이씨앤씨학원": "다중이용시설",
    "영통역통합": "다중이용시설",
    "좋은이웃데이케어센터통합": "노인요양시설"
}

standards_by_facility_type = {
    "다중이용시설": {
        "pm10": 100,
        "pm25": 50,
        "co2": 1000,
        "hcho": 100,
        "co": 10,
        "no2": 100,
        "voc": 500
    },
    "어린이집": {
        "pm10": 75,
        "pm25": 35,
        "co2": 1000,
        "voc": 400,
        "hcho": 80,
        "co": 10,
        "no2": 50
    },
    "노인요양시설": {
        "pm10": 75,
        "pm25": 35,
        "co2": 1000,
        "voc": 400,
        "hcho": 80,
        "co": 10,
        "no2": 50
    },
    "실내주차장": {
        "pm10": 200,
        "hcho": 100,
        "co": 25,
        "no2": 300,
        "co2": 1000,
        "voc": 1000
    },
    "실내체육시설": {
        "pm10": 200,
        "pm25": 50,
        "co2": 1000,
        "hcho": 100,
        "co": 10,
        "no2": 100,
        "voc": 500
    }
}

def get_facility_type(place_name: str) -> str:
    if "좋은이웃데이케어센터" in place_name and place_name not in facility_type_mapping:
        return facility_type_mapping.get("좋은이웃데이케어센터", None)
    return facility_type_mapping.get(place_name, None)

def get_standard_line(place_name: str, pollutant: str):
    ftype = get_facility_type(place_name)
    if not ftype:
        return None
    return standards_by_facility_type.get(ftype, {}).get(pollutant, None)

# -----------------------------------------------------------------------------
# (3) CSV 불러오기
# -----------------------------------------------------------------------------
ob1_min = pd.read_csv("C:/muf/input/가산A1타워주차장_20230331.csv", encoding='utf-8')
ob2_min = pd.read_csv("C:/muf/input/에이샵스크린골프_20230331.csv", encoding='utf-8')
ob31_min = pd.read_csv("C:/muf/input/영통역대합실_20230331.csv", encoding='utf-8')
ob32_min = pd.read_csv("C:/muf/input/영통역지하역사_20230331.csv", encoding='utf-8')
ob4_min = pd.read_csv("C:/muf/input/이든어린이집_20230331.csv", encoding='utf-8')
ob51_min = pd.read_csv("C:/muf/input/좋은이웃데이케어센터1_20230331.csv", encoding='utf-8')
ob52_min = pd.read_csv("C:/muf/input/좋은이웃데이케어센터2_20230331.csv", encoding='utf-8')
ob53_min = pd.read_csv("C:/muf/input/좋은이웃데이케어센터3_20230331.csv", encoding='utf-8')
ob54_min = pd.read_csv("C:/muf/input/좋은이웃데이케어센터4_20230331.csv", encoding='utf-8')
ob6_min = pd.read_csv("C:/muf/input/하이씨앤씨학원_20230331.csv", encoding='utf-8')

# -----------------------------------------------------------------------------
# (4) 여러 CSV -> (tmfc_d, tmfc_h) 통합
# -----------------------------------------------------------------------------
def rowwise_average_dataframes(dfs):
    if not dfs:
        return pd.DataFrame()
    common_cols = set(dfs[0].columns)
    for df in dfs[1:]:
        common_cols = common_cols.intersection(set(df.columns))
    common_cols = list(common_cols)
    if 'tmfc_d' not in common_cols or 'tmfc_h' not in common_cols:
        raise ValueError("DataFrames must have tmfc_d, tmfc_h columns.")

    numeric_cols = []
    for col in common_cols:
        if col not in ['tmfc_d', 'tmfc_h'] and pd.api.types.is_numeric_dtype(dfs[0][col]):
            numeric_cols.append(col)

    for df in dfs:
        df.sort_values(by=['tmfc_d', 'tmfc_h'], inplace=True, ignore_index=True)

    all_times = set()
    for df in dfs:
        for row in df[['tmfc_d','tmfc_h']].itertuples(index=False):
            all_times.add((row.tmfc_d, row.tmfc_h))
    all_times = sorted(all_times, key=lambda x:(x[0],x[1]))

    result_rows = []
    for (d,h) in all_times:
        sub_records_list = []
        for df in dfs:
            sub_df = df[(df['tmfc_d']==d)&(df['tmfc_h']==h)]
            sub_records_list.append(sub_df.to_dict('records'))
        max_len = max(len(rlist) for rlist in sub_records_list)
        for i in range(max_len):
            row_data = {'tmfc_d':d, 'tmfc_h':h}
            for col in numeric_cols:
                vals=[]
                for rlist in sub_records_list:
                    if i < len(rlist):
                        val=rlist[i].get(col, np.nan)
                        if pd.notnull(val):
                            vals.append(val)
                row_data[col] = np.mean(vals) if len(vals)>0 else np.nan
            result_rows.append(row_data)

    merged_df = pd.DataFrame(result_rows)
    merged_df.sort_values(by=['tmfc_d','tmfc_h'], inplace=True, ignore_index=True)
    return merged_df

# -----------------------------------------------------------------------------
# (5) 그래프 그리는 함수 (시계열 제외, 박스플롯만)
# -----------------------------------------------------------------------------
def plot_air_quality_for_location(
    df: pd.DataFrame,
    place_name: str,
    pollutants: list,
    pollutant_scale_dict: dict,
    save_dir: str
):
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    # 영문 장소명
    loc_english = location_mapping.get(place_name, place_name)

    # datetime 변환
    df['datetime'] = pd.to_datetime(df['tmfc_d'].astype(str)+df['tmfc_h'].astype(str).str.zfill(2),
                                    format='%Y%m%d%H')
    df['tmfc_h'] = df['tmfc_h'].astype(int)%24
    df['year_month'] = df['datetime'].dt.strftime('%Y-%m')

    # 박스플롯 전용 라벨 (LaTeX)
    pollutant_label_dict = {
        "pm10": r"PM$_{10}$ ($\mu g/m^3$)",
        "pm25": r"PM$_{2.5}$ ($\mu g/m^3$)",
        "co2": r"CO$_2$ (ppm)",
        "voc": r"VOC ($\mu g/m^3$)",
        "temp": "Temperature (°C)",
        "humi": "Humidity (%)",
        "hcho": r"HCHO ($\mu g/m^3$)",
        "co": r"CO (ppm)",
        "no2": r"NO$_2$ (ppb)"
    }

    for pollutant in pollutants:
        if pollutant not in df.columns:
            continue

        # y축 스케일
        scale_y = pollutant_scale_dict.get(pollutant, "linear")
        if scale_y in ["log","symlog"]:
            df.loc[df[pollutant]<=0, pollutant] = np.nan

        y_label = pollutant_label_dict.get(pollutant, pollutant_english.get(pollutant, pollutant))
        cut_conc = get_standard_line(place_name, pollutant)

        # 시간별 박스플롯
        plt.figure(figsize=(9,6))
        plt.title(f"{loc_english} {pollutant_english.get(pollutant,pollutant)} Hourly Boxplot", fontsize=14)
        hours_sorted = sorted(df["tmfc_h"].unique())
        grouped_data_hour = [ df.loc[df["tmfc_h"]==h, pollutant].dropna() for h in hours_sorted ]
        plt.boxplot(grouped_data_hour, positions=range(1, len(hours_sorted)+1))
        plt.xlabel("Hour of Day")
        plt.ylabel(y_label)
        plt.yscale(scale_y)
        plt.xticks(range(1, len(hours_sorted)+1), [str(h) for h in hours_sorted])

        if cut_conc is not None:
            plt.axhline(y=cut_conc, color='red', linestyle='--')

        means_hourly = [g.mean() for g in grouped_data_hour]
        plt.scatter(range(1, len(hours_sorted)+1), means_hourly, color='darkblue')

        plt.tight_layout()
        outfile = f"{save_dir}/hourly_boxplot_{loc_english}_{pollutant_english.get(pollutant,pollutant)}.png"
        plt.savefig(outfile, dpi=300)
        plt.close()

        # 월별 박스플롯
        df_nonan = df.dropna(subset=[pollutant])
        valid_year_months = sorted(df_nonan["year_month"].unique())
        if len(valid_year_months)==0:
            continue
        plt.figure(figsize=(9,6))
        plt.title(f"{loc_english} {pollutant_english.get(pollutant,pollutant)} Monthly Boxplot", fontsize=14)
        grouped_data_month = [ df_nonan.loc[df_nonan["year_month"]==ym, pollutant].dropna() for ym in valid_year_months ]
        plt.boxplot(grouped_data_month, positions=range(1,len(valid_year_months)+1))
        plt.xlabel("Year-Month")
        plt.ylabel(y_label)
        plt.yscale(scale_y)
        plt.xticks(range(1, len(valid_year_months)+1), valid_year_months, rotation=45)

        if cut_conc is not None:
            plt.axhline(y=cut_conc, color='red', linestyle='--')

        means_monthly = [arr.mean() for arr in grouped_data_month]
        plt.scatter(range(1, len(valid_year_months)+1), means_monthly, color='darkblue')

        plt.tight_layout()
        outfile = f"{save_dir}/monthly_boxplot_{loc_english}_{pollutant_english.get(pollutant,pollutant)}.png"
        plt.savefig(outfile, dpi=300)
        plt.close()

# -----------------------------------------------------------------------------
# (6) 초과 횟수/전체/초과율 계산 (전체 or 특정 월)
# -----------------------------------------------------------------------------
def calc_exceed_ratio(df, place_name, pollutant):
    """df에서 pollutant 유효데이터 중 기준을 초과(>)하는 개수 비율."""
    cut = get_standard_line(place_name, pollutant)
    valid_data = df[pollutant].dropna()
    total_count = len(valid_data)
    if cut is not None and total_count>0:
        exceed_count = sum(valid_data>cut)
        exceed_ratio = (exceed_count/total_count)*100.0
        return exceed_count, total_count, exceed_ratio
    else:
        return 0, total_count, 0.0

# -----------------------------------------------------------------------------
# (7) 엑셀에 기록하는 함수
# -----------------------------------------------------------------------------
def write_results_to_sheet(ws, results_dict):
    """
    results_dict 구조:
        { place_name(한글) : { pollutant : (exceed_count, total_count, ratio) }, ... }

    (수정 사항)
    - 물질 표시 순서는 고정 리스트 ["pm10","pm25","co2","voc","temp","humi","hcho","co","no2"] 순서를 따름
    - 1열(첫 번째 열)은 해당 행의 데이터 종류(예: '장소명(영문)', '물질', '초과횟수', '전체개수', '초과율')가 표시됨.
    - 장소 하나당 다음과 같은 형식으로 6행을 사용:
        1) row:   A열='장소명(영문)', B열=장소명
        2) row+1: A열='물질', B열부터=물질리스트
        3) row+2: A열='초과횟수', B열부터=초과횟수
        4) row+3: A열='전체개수', B열부터=전체개수
        5) row+4: A열='초과율',   B열부터=초과율
        6) row+5: 빈 행
    - 초과율>2%인 셀들은 배경을 연핑크색으로
    """
    pink_fill = PatternFill(start_color='FFFFC0CB', end_color='FFFFC0CB', fill_type='solid')

    # 고정된 물질 순서
    ordered_pollutants = ["pm10","pm25","co2","voc","temp","humi","hcho","co","no2"]

    current_row = 1
    for place_name, data_dict in results_dict.items():
        loc_english = location_mapping.get(place_name, place_name)

        # data_dict.keys() 중 실제 데이터가 있는 물질만 추려서, ordered_pollutants 순서대로 정렬
        pol_list = [p for p in ordered_pollutants if p in data_dict.keys()]

        # (1) 첫 행: 장소명(영문)
        ws.cell(row=current_row, column=1, value="장소명(영문)")
        ws.cell(row=current_row, column=2, value=loc_english)
        current_row += 1

        # (2) 물질 행
        ws.cell(row=current_row, column=1, value="물질")
        for c, pol in enumerate(pol_list, start=2):
            pol_eng = pollutant_english.get(pol, pol)
            cell = ws.cell(row=current_row, column=c)
            cell.value = pol_eng
        current_row += 1

        # (3) 초과횟수 행
        ws.cell(row=current_row, column=1, value="초과횟수")
        for c, pol in enumerate(pol_list, start=2):
            exceed_count, _, ratio = data_dict[pol]
            cell_exceed = ws.cell(row=current_row, column=c, value=exceed_count)
            # 초과율 판단해서 2% 초과 시 색칠(물질명/초과횟수/전체개수/초과율 모두)
            if ratio > 2.0:
                # 물질명 칸은 (current_row - 1)행, 같은 column
                ws.cell(row=current_row - 1, column=c).fill = pink_fill
                cell_exceed.fill = pink_fill
        current_row += 1

        # (4) 전체개수 행
        ws.cell(row=current_row, column=1, value="전체개수")
        for c, pol in enumerate(pol_list, start=2):
            _, total_count, ratio = data_dict[pol]
            cell_total = ws.cell(row=current_row, column=c, value=total_count)
            if ratio > 2.0:
                cell_total.fill = pink_fill
        current_row += 1

        # (5) 초과율 행
        ws.cell(row=current_row, column=1, value="초과율(%)")
        for c, pol in enumerate(pol_list, start=2):
            _, _, ratio = data_dict[pol]
            cell_ratio = ws.cell(row=current_row, column=c, value=ratio)
            if ratio > 2.0:
                cell_ratio.fill = pink_fill
        current_row += 1

        # (6) 빈 행
        current_row += 1

# -----------------------------------------------------------------------------
# (8) 실제 실행
# -----------------------------------------------------------------------------
if __name__=="__main__":
    # 1) 영통역 통합
    ob3_min = rowwise_average_dataframes([ob31_min, ob32_min])
    # 2) 좋은이웃데이케어센터 통합
    ob5_min = rowwise_average_dataframes([ob51_min, ob52_min, ob53_min, ob54_min])

    # 그래프용 오염물질 목록
    target_pollutants = ["pm10","pm25","co2","voc","temp","humi","hcho","co","no2"]

    pollutant_scale_dict = {
        "pm10": "symlog",
        "pm25": "symlog",
        "co2": "linear",
        "voc": "symlog",
        "temp": "linear",
        "humi": "linear",
        "hcho": "symlog",
        "co": "symlog",
        "no2": "symlog"
    }

    # 그래프 생성
    plot_air_quality_for_location(ob1_min, "가산A1타워주차장", target_pollutants, pollutant_scale_dict, "C:/muf/result/graph/1_지하주차장")
    plot_air_quality_for_location(ob2_min, "에이샵스크린골프", target_pollutants, pollutant_scale_dict, "C:/muf/result/graph/2_스크린골프장")
    plot_air_quality_for_location(ob3_min, "영통역통합",       target_pollutants, pollutant_scale_dict, "C:/muf/result/graph/3_지하철")
    plot_air_quality_for_location(ob4_min, "이든어린이집",     target_pollutants, pollutant_scale_dict, "C:/muf/result/graph/4_어린이집")
    plot_air_quality_for_location(ob5_min, "좋은이웃데이케어센터통합", target_pollutants, pollutant_scale_dict, "C:/muf/result/graph/5_노인요양시설")
    plot_air_quality_for_location(ob6_min, "하이씨앤씨학원",   target_pollutants, pollutant_scale_dict, "C:/muf/result/graph/6_학원")

    # -----------------------------------------------------------------------------
    # (A) 엑셀에 쓸 통계를 위한 준비
    # -----------------------------------------------------------------------------
    # 1) facility-DataFrame 목록
    facilities_data = [
        ("가산A1타워주차장", ob1_min),
        ("에이샵스크린골프", ob2_min),
        ("영통역통합", ob3_min),
        ("이든어린이집", ob4_min),
        ("좋은이웃데이케어센터통합", ob5_min),
        ("하이씨앤씨학원", ob6_min)
    ]

    # 2) 모든 df에 datetime / year_month 열 생성(월별 통계용)
    for (place_name, df) in facilities_data:
        df['datetime'] = pd.to_datetime(df['tmfc_d'].astype(str)+df['tmfc_h'].astype(str).str.zfill(2),
                                        format='%Y%m%d%H', errors='coerce')
        df['year_month'] = df['datetime'].dt.strftime('%Y-%m')

    # 3) 모든 월(YYYY-MM) 수집
    all_months = set()
    for (place_name, df) in facilities_data:
        all_months.update(df['year_month'].dropna().unique())
    all_months = sorted(list(all_months))  # 예: ['2023-01','2023-02',...]

    # -----------------------------------------------------------------------------
    # (B) 전체기간/월별 통계 계산 -> 파이썬 dict 에 저장
    # -----------------------------------------------------------------------------
    results_whole = {}
    results_monthly = defaultdict(dict)

    for (place_name, df) in facilities_data:
        # 전체기간 통계
        poll_dict = {}
        for p in target_pollutants:
            if p in df.columns:
                exc, tot, rat = calc_exceed_ratio(df, place_name, p)
                poll_dict[p] = (exc, tot, rat)
        results_whole[place_name] = poll_dict

        # 월별 통계
        for ym in df['year_month'].dropna().unique():
            sub_df = df[df['year_month']==ym]
            pol_dict_m = {}
            for p in target_pollutants:
                if p in sub_df.columns:
                    exc, tot, rat = calc_exceed_ratio(sub_df, place_name, p)
                    pol_dict_m[p] = (exc, tot, rat)
            if len(pol_dict_m)>0:
                results_monthly[ym][place_name] = pol_dict_m

    # -----------------------------------------------------------------------------
    # (C) openpyxl 기반으로 여러 시트에 작성
    # -----------------------------------------------------------------------------
    wb = Workbook()

    # 1) 첫 시트: 전체기간
    ws_whole = wb.active
    ws_whole.title = "전체기간"
    write_results_to_sheet(ws_whole, results_whole)

    # 2) 이후 시트들: 월별
    for ym in all_months:
        ws_m = wb.create_sheet(title=str(ym))
        monthly_dict = results_monthly[ym] if ym in results_monthly else {}
        if monthly_dict:
            write_results_to_sheet(ws_m, monthly_dict)
        else:
            pass  # 해당 월에 데이터가 아예 없으면 비워둠

    out_path = "C:/muf/result/exceedance_ratio.xlsx"
    wb.save(out_path)

    print(f"모든 그래프 생성 및 통계 엑셀({out_path}) 작성 완료!")
